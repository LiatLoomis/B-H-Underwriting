import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

# Load embedded BH template (from disk)
TEMPLATE_PATH = "BH_Underwriting_Template.xlsx"

st.set_page_config(page_title="BH Underwriting AI", layout="centered")
st.title("🏢 BH Property Underwriting Tool")

# Password
if "auth" not in st.session_state:
    pwd = st.text_input("Enter password to use the app:", type="password")
    if pwd != "1234":
        st.warning("Incorrect or missing password.")
        st.stop()
    else:
        st.session_state.auth = True

# Upload rent roll
st.header("📁 Upload Rent Roll (.xlsx)")
rent_roll_file = st.file_uploader("Upload Rent Roll", type=["xlsx"], help="Excel format only. Rent column can have any name.")

# Input fields
st.header("🏦 Property & Loan Info")
price = st.number_input("Purchase Price ($)", value=1000000)
loan = st.number_input("Loan Amount ($)", value=700000)
rate = st.number_input("Interest Rate (%)", value=6.0)
term = st.number_input("Loan Term (Years)", value=30)
vacancy = st.number_input("Vacancy Rate (%)", value=5.0)
lease_type = st.selectbox("Lease Type", ["Gross", "NNN"])

run = st.button("Run Underwriting & Download")

if run:
    if not rent_roll_file:
        st.error("Please upload a rent roll file.")
    else:
        df = pd.read_excel(rent_roll_file)

        # Try to find the rent column
        rent_col = None
        for col in df.columns:
            if "rent" in str(col).lower():
                rent_col = col
                break
        if rent_col is None:
            rent_col = df.columns[-1]  # fallback to last column

        rent_data = pd.to_numeric(df[rent_col], errors="coerce")
        total_monthly_rent = rent_data.sum(skipna=True)
        annual_income = total_monthly_rent * 12 * (1 - vacancy / 100)

        # Expenses
        if lease_type == "NNN":
            expenses = 0
        else:
            expenses = 0.012 * price + 8000 + 10000 + 6000

        noi = annual_income - expenses

        # Loan calculations
        r = rate / 100 / 12
        n = term * 12
        pmt = loan * r / (1 - (1 + r) ** -n)
        annual_debt = pmt * 12

        taxable = noi - annual_debt
        after_tax = taxable * (1 - 0.35)
        coc = after_tax / (price - loan)
        cap_rate = noi / price

        # Load and fill built-in BH template
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        ws["C24"] = annual_income
        ws["C31"] = expenses
        ws["C43"] = noi
        ws["C55"] = taxable
        ws["C56"] = after_tax
        ws["C61"] = coc
        ws["C62"] = cap_rate

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Underwriting completed!")
        st.download_button(
            label="📥 Download Completed Excel",
            data=output,
            file_name="BH_Underwriting.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
